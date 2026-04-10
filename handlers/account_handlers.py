from aiogram import Router, F
from aiogram.filters.command import Command
from aiogram.types import (
    Message,
    InlineKeyboardButton,
    InlineKeyboardMarkup,
    CallbackQuery,
    BufferedInputFile,
)
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import io
from datetime import datetime
from pathlib import Path

from core.config import API_HASH, API_ID
from core.state import app_state
from database import get_user_accounts
from services.session_service import ensure_connected_client
from services.user_paths import session_base_path

router = Router()
LOGIN_REQUIRED_TEXT = "\u274c \u0421\u043d\u0430\u0447\u0430\u043b\u0430 \u0432\u043e\u0439\u0434\u0438 \u0447\u0435\u0440\u0435\u0437 /login"

# user_id -> account_number
user_current_account = {}


def _status_text(is_active: bool) -> str:
    return "\U0001f7e2" if is_active else "\U0001f534"


def _h(text: str) -> str:
    return (text or "").replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")


def _get_session_age_days(user_id: int, account_number: int) -> int | None:
    candidates = [
        Path(f"{session_base_path(user_id, account_number)}.session"),
        Path(__file__).resolve().parent.parent / f"session_{user_id}_{account_number}.session",
    ]
    session_file = next((candidate for candidate in candidates if candidate.exists()), None)
    if not session_file:
        return None

    created_at = datetime.fromtimestamp(session_file.stat().st_mtime)
    return max((datetime.now() - created_at).days, 0)


def _format_session_age(user_id: int, account_number: int) -> str:
    age_days = _get_session_age_days(user_id, account_number)
    if age_days is None:
        return "\u043d\u0435\u0438\u0437\u0432\u0435\u0441\u0442\u043d\u043e"
    if age_days == 0:
        return "\u0441\u0435\u0433\u043e\u0434\u043d\u044f"
    return f"{age_days} \u0434\u043d."


async def show_accounts_menu(message: Message, user_id: int, edit: bool = False):
    accounts = get_user_accounts(user_id)
    if not accounts:
        text = LOGIN_REQUIRED_TEXT
        if edit:
            await message.edit_text(text)
        else:
            await message.answer(text)
        return

    info = "👤 <b>МОИ АККАУНТЫ</b>\n"
    info += f"📦 Всего: <b>{len(accounts)}</b>\n"
    info += "━━━━━━━━━━━━━━━━\n\n"
    buttons = []

    for account_number, telegram_id, username, first_name, is_active in accounts:
        status = _status_text(is_active)
        first_name_safe = _h(first_name or "")
        session_age = _format_session_age(user_id, account_number)
        info += f"{status} <b>{first_name_safe}</b> • #{account_number}\n"
        info += f"   \u23f3 \u0421\u0435\u0441\u0441\u0438\u044f: {session_age}\n"
        buttons.append(
            [
                InlineKeyboardButton(
                    text=f"{status} {first_name or ''} (#{account_number})",
                    callback_data=f"view_account_{account_number}",
                )
            ]
        )

    info += "\n━━━━━━━━━━━━━━━━"

    buttons.append(
        [InlineKeyboardButton(text="⬅️ Назад", callback_data="close_accounts_menu")]
    )
    inline_keyboard = InlineKeyboardMarkup(inline_keyboard=buttons)

    if edit:
        await message.edit_text(info, reply_markup=inline_keyboard, parse_mode="HTML")
    else:
        await message.answer(info, reply_markup=inline_keyboard, parse_mode="HTML")


@router.message(
    F.text.contains("\u041c\u043e\u0439 \u0430\u043a\u043a\u0430\u0443\u043d\u0442")
)
async def account_button(message: Message):
    user_id = message.from_user.id
    try:
        await message.delete()
    except Exception:
        pass
    await show_accounts_menu(message, user_id, edit=False)


@router.message(Command("menu"))
async def cmd_menu(message: Message):
    user_id = message.from_user.id
    await show_accounts_menu(message, user_id, edit=False)


@router.callback_query(F.data.startswith("view_account_"))
async def view_account(query: CallbackQuery):
    user_id = query.from_user.id
    account_number = int(query.data.split("_")[2])

    await query.answer()
    user_current_account[user_id] = account_number

    client = await ensure_connected_client(
        user_id,
        account_number,
        api_id=API_ID,
        api_hash=API_HASH,
    )
    if not client:
        accounts = get_user_accounts(user_id)
        account_info = next((acc for acc in accounts if acc[0] == account_number), None)

        if account_info:
            _, telegram_id, username, first_name, is_active = account_info
            status = _status_text(is_active)
            first_name_safe = _h(
                first_name
                or "\u041d\u0435\u0438\u0437\u0432\u0435\u0441\u0442\u043d\u043e"
            )
            username_safe = _h(
                username or "\u043d\u0435 \u0443\u043a\u0430\u0437\u0430\u043d"
            )
            session_age = _format_session_age(user_id, account_number)
            info = (
                "🧾 <b>ИНФОРМАЦИЯ ОБ АККАУНТЕ</b>\n"
                "━━━━━━━━━━━━━━━━\n\n"
                f"{status} <b>{first_name_safe}</b> • #{account_number}\n"
                f"🆔 ID: <code>{telegram_id}</code>\n"
                f"\U0001f517 Username: @{username_safe}\n"
                f"\u23f3 \u0421\u0435\u0441\u0441\u0438\u044f: <b>{session_age}</b>\n\n"
                f"{LOGIN_REQUIRED_TEXT}."
            )
            inline_keyboard = InlineKeyboardMarkup(
                inline_keyboard=[
                    [
                        InlineKeyboardButton(
                            text="⬅️ Назад", callback_data="back_to_account_menu"
                        )
                    ]
                ]
            )
            await query.message.edit_text(
                info, reply_markup=inline_keyboard, parse_mode="HTML"
            )
        else:
            await query.message.edit_text(
                "\u0410\u043a\u043a\u0430\u0443\u043d\u0442 \u043d\u0435 \u043d\u0430\u0439\u0434\u0435\u043d. \u0418\u0441\u043f\u043e\u043b\u044c\u0437\u0443\u0439 /login \u0434\u043b\u044f \u0432\u0445\u043e\u0434\u0430."
            )
        return

    await refresh_menu_content(
        query.message, query, user_id, account_number, is_refresh=False
    )


@router.callback_query(F.data == "get_chats_list")
async def show_chats_list_selection(query: CallbackQuery):
    await query.answer()

    keyboard = InlineKeyboardMarkup(
        inline_keyboard=[
            [
                InlineKeyboardButton(
                    text="\u041b\u0438\u0447\u043d\u044b\u0435 \u0447\u0430\u0442\u044b",
                    callback_data="export_private_chats",
                )
            ],
            [
                InlineKeyboardButton(
                    text="\u0413\u0440\u0443\u043f\u043f\u044b",
                    callback_data="export_groups",
                )
            ],
            [
                InlineKeyboardButton(
                    text="\u041a\u0430\u043d\u0430\u043b\u044b",
                    callback_data="export_channels",
                )
            ],
            [
                InlineKeyboardButton(
                    text="\u0412\u0441\u0435 \u0441\u0440\u0430\u0437\u0443",
                    callback_data="export_all_chats",
                )
            ],
            [
                InlineKeyboardButton(
                    text="\u2b05\ufe0f \u041d\u0430\u0437\u0430\u0434",
                    callback_data="back_to_account_menu",
                )
            ],
        ]
    )

    await query.message.edit_text(
        "<b>\u0412\u044b\u0431\u0435\u0440\u0438\u0442\u0435, \u0447\u0442\u043e \u044d\u043a\u0441\u043f\u043e\u0440\u0442\u0438\u0440\u043e\u0432\u0430\u0442\u044c \u0432 Excel:</b>\n\n"
        "\u2022 <b>\u041b\u0438\u0447\u043d\u044b\u0435 \u0447\u0430\u0442\u044b</b> - \u0442\u043e\u043b\u044c\u043a\u043e DM \u0438 \u0431\u043e\u0442\u044b\n"
        "\u2022 <b>\u0413\u0440\u0443\u043f\u043f\u044b</b> - \u0442\u043e\u043b\u044c\u043a\u043e \u0433\u0440\u0443\u043f\u043f\u044b\n"
        "\u2022 <b>\u041a\u0430\u043d\u0430\u043b\u044b</b> - \u0442\u043e\u043b\u044c\u043a\u043e \u043a\u0430\u043d\u0430\u043b\u044b\n"
        "\u2022 <b>\u0412\u0441\u0435 \u0441\u0440\u0430\u0437\u0443</b> - \u0432\u0441\u0435 \u0447\u0430\u0442\u044b \u0441 \u043e\u0442\u0434\u0435\u043b\u044c\u043d\u044b\u043c\u0438 \u0442\u0430\u0431\u043b\u0438\u0446\u0430\u043c\u0438",
        reply_markup=keyboard,
        parse_mode="HTML",
    )


async def export_chats_by_type(query: CallbackQuery, chat_types: list):
    await query.answer(
        "\u041f\u043e\u043b\u043d\u044b\u0439 \u0441\u043a\u0430\u043d \u0430\u043a\u043a\u0430\u0443\u043d\u0442\u0430, \u043c\u043e\u0436\u0435\u0442 \u0431\u044b\u0442\u044c \u0434\u043e\u043b\u0433\u043e...",
        show_alert=False,
    )
    user_id = query.from_user.id

    try:
        account_number = user_current_account.get(user_id)
        if not account_number:
            await query.answer(
                "\u0410\u043a\u043a\u0430\u0443\u043d\u0442 \u043d\u0435 \u0432\u044b\u0431\u0440\u0430\u043d",
                show_alert=True,
            )
            return

        client = await ensure_connected_client(
            user_id,
            account_number,
            api_id=API_ID,
            api_hash=API_HASH,
        )
        if not client:
            await query.answer(
                "\u0410\u043a\u043a\u0430\u0443\u043d\u0442 \u043d\u0435 \u043f\u043e\u0434\u043a\u043b\u044e\u0447\u0435\u043d",
                show_alert=True,
            )
            return

        me = await client.get_me()

        all_dialogs = []
        async for dialog in client.iter_dialogs(limit=None):
            all_dialogs.append(dialog)

        private_chats = []
        groups = []
        channels = []
        bots = []
        my_groups = []
        my_channels = []

        from telethon.types import User, Chat, Channel

        for dialog in all_dialogs:
            entity = dialog.entity
            name = dialog.name
            username = None
            owner_name = None

            if hasattr(entity, "username") and entity.username:
                username = entity.username

            try:
                if isinstance(entity, User):
                    owner_name = entity.first_name or ""
                    if entity.bot:
                        bots.append((name, dialog.id, username, owner_name))
                    else:
                        private_chats.append((name, dialog.id, username, owner_name))
                elif isinstance(entity, Channel):
                    is_mega = getattr(entity, "megagroup", False)
                    is_group = getattr(entity, "is_group", False)
                    is_creator = getattr(entity, "is_creator", False)

                    try:
                        admin_info = await client.get_permissions(dialog, me)
                        if hasattr(admin_info, "creator") and admin_info.creator:
                            owner_name = "\u042f \u0441\u043e\u0437\u0434\u0430\u0442\u0435\u043b\u044c"
                    except Exception:
                        pass

                    if is_mega or is_group:
                        if is_creator:
                            my_groups.append(
                                (name, dialog.id, username, owner_name or "")
                            )
                        else:
                            groups.append((name, dialog.id, username, owner_name or ""))
                    else:
                        if is_creator:
                            my_channels.append(
                                (name, dialog.id, username, owner_name or "")
                            )
                        else:
                            channels.append(
                                (name, dialog.id, username, owner_name or "")
                            )
                elif isinstance(entity, Chat):
                    creator_owner = ""
                    if hasattr(entity, "creator_id") and entity.creator_id == me.id:
                        creator_owner = "\u042f \u0441\u043e\u0437\u0434\u0430\u0442\u0435\u043b\u044c"
                        my_groups.append((name, dialog.id, username, creator_owner))
                    else:
                        groups.append((name, dialog.id, username, creator_owner))
                else:
                    private_chats.append((name, dialog.id, username, owner_name or ""))
            except Exception:
                private_chats.append((name, dialog.id, username, ""))

        wb = Workbook()
        wb.remove(wb.active)

        header_fill = PatternFill(
            start_color="4472C4", end_color="4472C4", fill_type="solid"
        )
        header_font = Font(bold=True, color="FFFFFF")

        if "all" in chat_types:
            sheets_data = [
                (
                    "\u041b\u0438\u0447\u043d\u044b\u0435 \u0447\u0430\u0442\u044b",
                    private_chats,
                ),
                ("\u0413\u0440\u0443\u043f\u043f\u044b", groups),
                ("\u041a\u0430\u043d\u0430\u043b\u044b", channels),
                ("\u0411\u043e\u0442\u044b", bots),
                ("\u041c\u043e\u0438 \u0433\u0440\u0443\u043f\u043f\u044b", my_groups),
                (
                    "\u041c\u043e\u0438 \u043a\u0430\u043d\u0430\u043b\u044b",
                    my_channels,
                ),
            ]
        else:
            sheets_data = []
            if "private" in chat_types:
                sheets_data.append(
                    (
                        "\u041b\u0438\u0447\u043d\u044b\u0435 \u0447\u0430\u0442\u044b",
                        private_chats,
                    )
                )
                sheets_data.append(("\u0411\u043e\u0442\u044b", bots))
            if "groups" in chat_types:
                sheets_data.append(("\u0413\u0440\u0443\u043f\u043f\u044b", groups))
                sheets_data.append(
                    ("\u041c\u043e\u0438 \u0433\u0440\u043f\u043f\u044b", my_groups)
                )
            if "channels" in chat_types:
                sheets_data.append(("\u041a\u0430\u043d\u0430\u043b\u044b", channels))
                sheets_data.append(
                    (
                        "\u041c\u043e\u0438 \u043a\u0430\u043d\u0430\u043b\u044b",
                        my_channels,
                    )
                )

        for sheet_name, chats_list in sheets_data:
            if not chats_list:
                continue
            ws = wb.create_sheet(sheet_name)
            ws["A1"] = "\u2116"
            ws["B1"] = "\u0418\u043c\u044f"
            ws["C1"] = "ID"
            ws["D1"] = "Username"
            ws["E1"] = (
                "\u0412\u043b\u0430\u0434\u0435\u043b\u0435\u0446/\u0422\u0438\u043f"
            )

            for cell in ["A1", "B1", "C1", "D1", "E1"]:
                ws[cell].fill = header_fill
                ws[cell].font = header_font
                ws[cell].alignment = Alignment(horizontal="center")

            for idx, (name, chat_id, username, owner_name) in enumerate(chats_list, 1):
                ws[f"A{idx + 1}"] = idx
                ws[f"B{idx + 1}"] = name
                ws[f"C{idx + 1}"] = chat_id
                ws[f"D{idx + 1}"] = username or ""
                ws[f"E{idx + 1}"] = owner_name or ""

            ws.column_dimensions["A"].width = 5
            ws.column_dimensions["B"].width = 30
            ws.column_dimensions["C"].width = 20
            ws.column_dimensions["D"].width = 20
            ws.column_dimensions["E"].width = 20

        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_bytes = excel_file.getvalue()

        await query.message.answer_document(
            BufferedInputFile(excel_bytes, filename="chats_list.xlsx"),
            caption="\u042d\u043a\u0441\u043f\u043e\u0440\u0442 \u0447\u0430\u0442\u043e\u0432",
        )
    except Exception as e:
        await query.answer(
            f"\u041e\u0448\u0438\u0431\u043a\u0430: {str(e)}", show_alert=True
        )


@router.callback_query(F.data == "export_private_chats")
async def export_private_chats(query: CallbackQuery):
    await export_chats_by_type(query, ["private"])


@router.callback_query(F.data == "export_groups")
async def export_groups(query: CallbackQuery):
    await export_chats_by_type(query, ["groups"])


@router.callback_query(F.data == "export_channels")
async def export_channels(query: CallbackQuery):
    await export_chats_by_type(query, ["channels"])


@router.callback_query(F.data == "export_all_chats")
async def export_all_chats(query: CallbackQuery):
    await export_chats_by_type(query, ["all"])


@router.callback_query(F.data == "back_to_account_menu")
async def back_to_account_menu(query: CallbackQuery):
    user_id = query.from_user.id
    await query.answer()
    await show_accounts_menu(query.message, user_id, edit=True)


@router.callback_query(F.data == "refresh_menu")
async def refresh_menu(query: CallbackQuery):
    user_id = query.from_user.id
    account_number = user_current_account.get(user_id)

    if not account_number:
        await query.answer(
            "\u0410\u043a\u043a\u0430\u0443\u043d\u0442 \u043d\u0435 \u0432\u044b\u0431\u0440\u0430\u043d",
            show_alert=True,
        )
        return

    await refresh_menu_content(
        query.message, query, user_id, account_number, is_refresh=True
    )


@router.callback_query(F.data == "close_accounts_menu")
async def close_accounts_menu_callback(query: CallbackQuery):
    await query.answer()
    try:
        await query.message.delete()
    except Exception:
        pass


async def refresh_menu_content(
    message: Message,
    query: CallbackQuery = None,
    user_id: int = None,
    account_number: int = None,
    is_refresh: bool = False,
):
    if user_id is None:
        user_id = message.from_user.id

    accounts = get_user_accounts(user_id)
    acc_info = next((acc for acc in accounts if acc[0] == account_number), None)
    if acc_info is not None and not acc_info[4]:
        if query:
            await query.answer(
                "\u0410\u043a\u043a\u0430\u0443\u043d\u0442 \u043e\u0442\u043a\u043b\u044e\u0447\u0435\u043d",
                show_alert=True,
            )
        else:
            await message.answer(
                "\u0410\u043a\u043a\u0430\u0443\u043d\u0442 \u043e\u0442\u043a\u043b\u044e\u0447\u0435\u043d"
            )
        return

    if user_id not in app_state.user_authenticated or not app_state.user_authenticated[user_id]:
        if query:
            await query.answer(LOGIN_REQUIRED_TEXT, show_alert=True)
        else:
            await message.answer(LOGIN_REQUIRED_TEXT)
        return

    client = await ensure_connected_client(
        user_id,
        account_number,
        api_id=API_ID,
        api_hash=API_HASH,
    )
    if not client:
        if query:
            await query.answer(
                "\u0410\u043a\u043a\u0430\u0443\u043d\u0442 \u043d\u0435 \u043f\u043e\u0434\u043a\u043b\u044e\u0447\u0435\u043d",
                show_alert=True,
            )
        else:
            await message.answer(
                "\u0410\u043a\u043a\u0430\u0443\u043d\u0442 \u043d\u0435 \u043f\u043e\u0434\u043a\u043b\u044e\u0447\u0435\u043d"
            )
        return

    try:
        me = await client.get_me()
        session_age = _format_session_age(user_id, account_number)

        info = (
            "👤 <b>МОЙ АККАУНТ</b>\n"
            "━━━━━━━━━━━━━━━━\n"
            f"📛 Имя: <b>{_h(me.first_name)}</b>\n"
            f"🆔 ID: <code>{me.id}</code>\n"
            f"🔗 Username: @{_h(me.username) if me.username else 'не указано'}\n"
            f"\U0001f4f1 \u041d\u043e\u043c\u0435\u0440: <code>{_h(me.phone)}</code>\n"
            f"\u23f3 \u0421\u0435\u0441\u0441\u0438\u044f: <b>{session_age}</b>\n\n"
            "ℹ️ Для полного списка и статистики используй кнопку «Список чатов»."
        )

        inline_keyboard = InlineKeyboardMarkup(
            inline_keyboard=[
                [
                    InlineKeyboardButton(
                        text="📋 Список чатов", callback_data="get_chats_list"
                    ),
                    InlineKeyboardButton(
                        text="🔄 Обновить", callback_data="refresh_menu"
                    ),
                ],
                [
                    InlineKeyboardButton(
                        text="⬅️ Назад", callback_data="back_to_account_menu"
                    )
                ],
            ]
        )

        if query:
            try:
                await message.edit_text(
                    info,
                    reply_markup=inline_keyboard,
                    parse_mode="HTML",
                    disable_web_page_preview=True,
                )
                if is_refresh:
                    notification = "Профиль обновлен"
                else:
                    notification = "\u0410\u043a\u043a\u0430\u0443\u043d\u0442 \u0432\u044b\u0431\u0440\u0430\u043d"
                await query.answer(notification, show_alert=False)
            except Exception as e:
                err = str(e).lower()
                if "not modified" in err:
                    await query.answer(
                        "\u0418\u043d\u0444\u043e\u0440\u043c\u0430\u0446\u0438\u044f \u043d\u0435 \u0438\u0437\u043c\u0435\u043d\u0438\u043b\u0430\u0441\u044c",
                        show_alert=False,
                    )
                    return
                if "can't parse entities" in err:
                    try:
                        await message.edit_text(
                            info,
                            reply_markup=inline_keyboard,
                            parse_mode=None,
                            disable_web_page_preview=True,
                        )
                        await query.answer(
                            "\u041e\u0431\u043d\u043e\u0432\u043b\u0435\u043d\u043e \u0431\u0435\u0437 \u0444\u043e\u0440\u043c\u0430\u0442\u0438\u0440\u043e\u0432\u0430\u043d\u0438\u044f",
                            show_alert=False,
                        )
                        return
                    except Exception:
                        pass
                if (
                    "message to edit not found" in err
                    or "message can't be edited" in err
                ):
                    try:
                        await message.answer(
                            info,
                            reply_markup=inline_keyboard,
                            parse_mode="HTML",
                            disable_web_page_preview=True,
                        )
                    except Exception:
                        pass
                    await query.answer(
                        "\u041e\u0442\u043f\u0440\u0430\u0432\u0438\u043b \u043d\u043e\u0432\u043e\u0435 \u0441\u043e\u043e\u0431\u0449\u0435\u043d\u0438\u0435.",
                        show_alert=False,
                    )
                    return
                await query.answer(
                    "\u041d\u0435 \u0443\u0434\u0430\u043b\u043e\u0441\u044c \u043e\u0431\u043d\u043e\u0432\u0438\u0442\u044c \u0441\u043e\u043e\u0431\u0449\u0435\u043d\u0438\u0435.",
                    show_alert=False,
                )
        else:
            await message.answer(
                info,
                reply_markup=inline_keyboard,
                parse_mode="HTML",
                disable_web_page_preview=True,
            )
    except Exception as e:
        if query:
            await query.answer(
                f"\u041e\u0448\u0438\u0431\u043a\u0430: {str(e)}", show_alert=True
            )
        else:
            await message.answer(
                f"\u041e\u0448\u0438\u0431\u043a\u0430 \u043f\u0440\u0438 \u0437\u0430\u0433\u0440\u0443\u0437\u043a\u0435 \u043c\u0435\u043d\u044e: {str(e)}"
            )
