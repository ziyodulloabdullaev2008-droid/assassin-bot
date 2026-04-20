# -*- coding: utf-8 -*-
from aiogram import F, Router
from aiogram.filters.command import Command
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.types import (
    BufferedInputFile,
    CallbackQuery,
    InlineKeyboardButton,
    InlineKeyboardMarkup,
    Message,
)
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

import io
import time
from html import escape

from core.config import ADMIN_ID
from database import (
    add_vip_user,
    get_all_users,
    get_all_vip_users,
    get_all_vip_users_with_details,
    get_user_accounts,
    is_vip_user,
    remove_vip_user,
    set_vip_session_limit,
)
from services.vip_service import (
    add_vip_user_to_cache,
    get_vip_cache_size,
    remove_vip_user_from_cache,
)

router = Router()


class VIPAddState(StatesGroup):
    waiting_for_user_id = State()
    waiting_for_duration = State()
    waiting_for_session_limit = State()
    waiting_for_edit_limit = State()


def _is_admin(user_id: int) -> bool:
    return user_id == ADMIN_ID


async def _reject_admin_message(message: Message) -> bool:
    if _is_admin(message.from_user.id):
        return False
    await message.answer("? Только владелец может управлять VIP")
    return True


async def _reject_admin_query(query: CallbackQuery) -> bool:
    if _is_admin(query.from_user.id):
        return False
    await query.answer("Только владелец может управлять VIP", show_alert=True)
    return True


def _build_users_lookup() -> dict[int, tuple[str, str]]:
    return {
        user_id: (username or "", first_name or "")
        for user_id, username, first_name, _ in get_all_users()
    }


def _format_username(username: str) -> str:
    if not username:
        return "без @username"
    return f"@{username.lstrip('@')}"


def _format_vip_time_left(expires_at: float | None) -> str:
    if expires_at is None:
        return "?"

    seconds_left = max(int(expires_at - time.time()), 0)
    days = seconds_left // 86400
    hours = (seconds_left % 86400) // 3600
    minutes = (seconds_left % 3600) // 60

    if days > 0:
        return f"{days} дн. {hours} ч."
    if hours > 0:
        return f"{hours} ч. {minutes} мин."
    return f"{minutes} мин."


def _format_session_limit(session_limit: int | None) -> str:
    if session_limit is None or int(session_limit) <= 0:
        return "?"
    return str(int(session_limit))


def _format_vip_entry(
    user_id: int,
    expires_at: float | None,
    session_limit: int | None,
    users_lookup: dict[int, tuple[str, str]],
) -> str:
    username, first_name = users_lookup.get(user_id, ("", ""))
    display_name = escape(first_name) if first_name else "без имени"
    display_username = escape(_format_username(username))
    left = escape(_format_vip_time_left(expires_at))
    limit_text = escape(_format_session_limit(session_limit))
    return (
        f"<code>{user_id}</code> - {display_name} - {display_username} - "
        f"{left} - лимит: <b>{limit_text}</b>"
    )


def _build_vip_list_text() -> str:
    vip_list = get_all_vip_users_with_details()
    users_lookup = _build_users_lookup()
    text = f"?? <b>VIP юзеры:</b> ({len(vip_list)})\n\n"
    for idx, (user_id, expires_at, session_limit) in enumerate(vip_list, 1):
        text += f"{idx}. {_format_vip_entry(user_id, expires_at, session_limit, users_lookup)}\n"
    return text


def _build_vips_overview_keyboard() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(
        inline_keyboard=[
            [InlineKeyboardButton(text="? Добавить", callback_data="vip_add_menu")],
            [InlineKeyboardButton(text="??? Лимиты", callback_data="vip_limits_menu")],
            [InlineKeyboardButton(text="? Удалить", callback_data="vip_delete_menu")],
        ]
    )


def _build_vip_delete_keyboard() -> InlineKeyboardMarkup:
    users_lookup = _build_users_lookup()
    rows = []
    for uid in get_all_vip_users():
        username, first_name = users_lookup.get(uid, ("", ""))
        label_parts = [str(uid)]
        if first_name:
            label_parts.append(first_name[:18])
        elif username:
            label_parts.append(_format_username(username)[:18])
        rows.append(
            [
                InlineKeyboardButton(
                    text="? " + " | ".join(label_parts),
                    callback_data=f"vip_remove_{uid}",
                )
            ]
        )

    rows.append([InlineKeyboardButton(text="?? Назад", callback_data="vip_back")])
    return InlineKeyboardMarkup(inline_keyboard=rows)


def _build_vip_limit_keyboard() -> InlineKeyboardMarkup:
    users_lookup = _build_users_lookup()
    rows = []
    for user_id, _, session_limit in get_all_vip_users_with_details():
        username, first_name = users_lookup.get(user_id, ("", ""))
        label = first_name or _format_username(username) or str(user_id)
        rows.append(
            [
                InlineKeyboardButton(
                    text=f"??? {user_id} | {label[:16]} | {_format_session_limit(session_limit)}",
                    callback_data=f"vip_limit_{user_id}",
                )
            ]
        )
    rows.append([InlineKeyboardButton(text="?? Назад", callback_data="vip_back")])
    return InlineKeyboardMarkup(inline_keyboard=rows)


def _vip_cancel_keyboard() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(
        inline_keyboard=[
            [InlineKeyboardButton(text="? Отменить", callback_data="vip_add_cancel")]
        ]
    )


def _vip_duration_prompt(user_id: int) -> str:
    return (
        "? <b>На какой срок выдать VIP?</b>\n\n"
        f"Юзер: <code>{user_id}</code>\n\n"
        "Введи число дней:\n"
        "• <code>0</code> — навсегда\n"
        "• <code>1</code> — на 1 день\n"
        "• <code>7</code> — на 7 дней"
    )


def _vip_limit_prompt(user_id: int, days: int) -> str:
    left = "?" if days == 0 else f"{days} дн."
    return (
        "?? <b>Какой лимит сессий поставить?</b>\n\n"
        f"Юзер: <code>{user_id}</code>\n"
        f"VIP: <b>{left}</b>\n\n"
        "Введи число:\n"
        "• <code>0</code> — без лимита\n"
        "• <code>1</code> — только 1 сессия\n"
        "• <code>3</code> — до 3 сессий"
    )


def _vip_edit_limit_prompt(user_id: int, current_limit: int | None) -> str:
    return (
        "?? <b>Изменить лимит сессий</b>\n\n"
        f"Юзер: <code>{user_id}</code>\n"
        f"Текущий лимит: <b>{_format_session_limit(current_limit)}</b>\n\n"
        "Введи новое число:\n"
        "• <code>0</code> — без лимита\n"
        "• <code>1</code>, <code>2</code>, <code>3</code> ..."
    )


async def _send_vip_duration_prompt(message: Message, state: FSMContext, user_id: int):
    await state.update_data(vip_user_id=user_id)
    await state.set_state(VIPAddState.waiting_for_duration)
    await message.answer(
        _vip_duration_prompt(user_id),
        parse_mode="HTML",
        reply_markup=_vip_cancel_keyboard(),
    )


async def _send_vip_limit_prompt(message: Message, state: FSMContext, days: int):
    data = await state.get_data()
    user_id = data.get("vip_user_id")
    if not user_id:
        await state.clear()
        await message.answer("? Не найден ID юзера. Начни заново: /vip <user_id>")
        return

    await state.update_data(vip_days=days)
    await state.set_state(VIPAddState.waiting_for_session_limit)
    await message.answer(
        _vip_limit_prompt(int(user_id), days),
        parse_mode="HTML",
        reply_markup=_vip_cancel_keyboard(),
    )


async def _finish_vip_add(message: Message, state: FSMContext, days: int, session_limit: int):
    data = await state.get_data()
    user_id = data.get("vip_user_id")
    if not user_id:
        await state.clear()
        await message.answer("? Не найден ID юзера. Начни заново: /vip <user_id>")
        return

    if add_vip_user(int(user_id), days, session_limit=session_limit):
        add_vip_user_to_cache(int(user_id))
        left = "?" if days == 0 else f"{days} дн."
        limit_text = _format_session_limit(session_limit)
        await message.answer(
            f"? VIP выдан юзеру <code>{user_id}</code>\n"
            f"Срок: <b>{left}</b>\n"
            f"Лимит сессий: <b>{limit_text}</b>",
            parse_mode="HTML",
        )
    else:
        await message.answer("? Не удалось выдать VIP")
    await state.clear()


@router.message(Command("vip"))
async def cmd_add_vip(message: Message, state: FSMContext):
    if await _reject_admin_message(message):
        return

    args = message.text.split()
    if len(args) < 2:
        await message.answer("? Использование: /vip <user_id>")
        return

    try:
        user_id = int(args[1])
    except ValueError:
        await message.answer("? Неверный ID")
        return

    await _send_vip_duration_prompt(message, state, user_id)


@router.message(Command("dlvip"))
async def cmd_remove_vip(message: Message):
    if await _reject_admin_message(message):
        return

    args = message.text.split()
    if len(args) < 2:
        await message.answer("? Использование: /dlvip <user_id>")
        return

    try:
        user_id = int(args[1])
    except ValueError:
        await message.answer("? Неверный ID")
        return

    if remove_vip_user(user_id):
        remove_vip_user_from_cache(user_id)
        await message.answer(f"? Юзер {user_id} удален из VIP")
    else:
        await message.answer(f"? Юзер {user_id} не найден в VIP")


@router.message(Command("vips"))
async def cmd_show_vips(message: Message):
    if await _reject_admin_message(message):
        return

    vip_list = get_all_vip_users()
    if not vip_list:
        await message.answer("?? VIP список пуст", reply_markup=_build_vips_overview_keyboard())
        return

    await message.answer(
        _build_vip_list_text(),
        parse_mode="HTML",
        reply_markup=_build_vips_overview_keyboard(),
    )


@router.callback_query(F.data == "vip_add_menu")
async def vip_add_menu(query: CallbackQuery, state: FSMContext):
    if await _reject_admin_query(query):
        return

    await query.answer()
    await state.set_state(VIPAddState.waiting_for_user_id)
    await query.message.edit_text(
        "?? <b>Введи ID пользователя для добавления:</b>\n(отправь число)",
        parse_mode="HTML",
        reply_markup=_vip_cancel_keyboard(),
    )


@router.callback_query(F.data == "vip_delete_menu")
async def vip_delete_menu(query: CallbackQuery):
    if await _reject_admin_query(query):
        return

    vip_list = get_all_vip_users()
    if not vip_list:
        await query.answer("VIP список пуст", show_alert=True)
        return

    await query.answer()
    await query.message.edit_text(
        "?? <b>Выберите VIP для удаления:</b>",
        parse_mode="HTML",
        reply_markup=_build_vip_delete_keyboard(),
    )


@router.callback_query(F.data == "vip_limits_menu")
async def vip_limits_menu(query: CallbackQuery):
    if await _reject_admin_query(query):
        return

    vip_list = get_all_vip_users()
    if not vip_list:
        await query.answer("VIP список пуст", show_alert=True)
        return

    await query.answer()
    await query.message.edit_text(
        "??? <b>Выберите VIP для изменения лимита:</b>",
        parse_mode="HTML",
        reply_markup=_build_vip_limit_keyboard(),
    )


@router.callback_query(F.data.startswith("vip_remove_"))
async def vip_remove_callback(query: CallbackQuery):
    if await _reject_admin_query(query):
        return

    user_id = int(query.data.split("_")[-1])
    if remove_vip_user(user_id):
        remove_vip_user_from_cache(user_id)
        await query.answer(f"? Юзер {user_id} удален из VIP")
        await vip_delete_menu(query)
    else:
        await query.answer("? Ошибка удаления", show_alert=True)


@router.callback_query(F.data.startswith("vip_limit_"))
async def vip_limit_callback(query: CallbackQuery, state: FSMContext):
    if await _reject_admin_query(query):
        return

    user_id = int(query.data.split("_")[-1])
    current_limit = None
    for vip_user_id, _, session_limit in get_all_vip_users_with_details():
        if vip_user_id == user_id:
            current_limit = session_limit
            break

    await state.update_data(vip_user_id=user_id)
    await state.set_state(VIPAddState.waiting_for_edit_limit)
    await query.answer()
    await query.message.edit_text(
        _vip_edit_limit_prompt(user_id, current_limit),
        parse_mode="HTML",
        reply_markup=_vip_cancel_keyboard(),
    )


@router.callback_query(F.data == "vip_back")
async def vip_back(query: CallbackQuery):
    if await _reject_admin_query(query):
        return

    await query.answer()
    await query.message.edit_text(
        _build_vip_list_text() if get_all_vip_users() else "?? VIP список пуст",
        parse_mode="HTML",
        reply_markup=_build_vips_overview_keyboard(),
    )


@router.callback_query(F.data == "vip_add_cancel")
async def vip_add_cancel(query: CallbackQuery, state: FSMContext):
    if await _reject_admin_query(query):
        return

    await state.clear()
    await query.answer("Отменено", show_alert=False)
    try:
        await query.message.edit_text("? Выдача/редактирование VIP отменены")
    except Exception:
        pass


@router.message(VIPAddState.waiting_for_user_id)
async def process_vip_user_id(message: Message, state: FSMContext):
    if await _reject_admin_message(message):
        await state.clear()
        return

    try:
        user_id = int(message.text.strip())
    except ValueError:
        await message.answer("? Неверный ID. Пожалуйста, введи число")
        return

    await _send_vip_duration_prompt(message, state, user_id)


@router.message(VIPAddState.waiting_for_duration)
async def process_vip_duration(message: Message, state: FSMContext):
    if await _reject_admin_message(message):
        await state.clear()
        return

    try:
        days = int(message.text.strip())
    except ValueError:
        await message.answer("? Введи число дней. Например: 0, 1, 7, 30")
        return

    if days < 0:
        await message.answer("? Срок не может быть меньше 0")
        return

    await _send_vip_limit_prompt(message, state, days)


@router.message(VIPAddState.waiting_for_session_limit)
async def process_vip_session_limit(message: Message, state: FSMContext):
    if await _reject_admin_message(message):
        await state.clear()
        return

    try:
        session_limit = int(message.text.strip())
    except ValueError:
        await message.answer("? Введи число. Например: 0, 1, 3")
        return

    if session_limit < 0:
        await message.answer("? Лимит не может быть меньше 0")
        return

    data = await state.get_data()
    days = int(data.get("vip_days", 0) or 0)
    await _finish_vip_add(message, state, days, session_limit)


@router.message(VIPAddState.waiting_for_edit_limit)
async def process_vip_edit_limit(message: Message, state: FSMContext):
    if await _reject_admin_message(message):
        await state.clear()
        return

    try:
        session_limit = int(message.text.strip())
    except ValueError:
        await message.answer("? Введи число. Например: 0, 1, 3")
        return

    if session_limit < 0:
        await message.answer("? Лимит не может быть меньше 0")
        return

    data = await state.get_data()
    user_id = data.get("vip_user_id")
    if not user_id:
        await state.clear()
        await message.answer("? Потерян ID пользователя. Открой /vips заново")
        return

    if set_vip_session_limit(int(user_id), session_limit):
        await message.answer(
            f"? Лимит сессий обновлен для <code>{user_id}</code>\n"
            f"Новый лимит: <b>{_format_session_limit(session_limit)}</b>",
            parse_mode="HTML",
        )
    else:
        await message.answer("? Не удалось обновить лимит")
    await state.clear()


@router.message(Command("users"))
async def cmd_show_users_stats(message: Message):
    if await _reject_admin_message(message):
        return

    users = get_all_users()
    vip_list = get_all_vip_users()
    vip_count = len(vip_list)

    stats_text = (
        "?? <b>Статистика бота:</b>\n\n"
        f"?? <b>Всего юзеров:</b> {len(users)}\n"
        f"?? <b>VIP юзеров:</b> {vip_count}\n"
        f"?? <b>Обычных юзеров:</b> {len(users) - vip_count}\n"
    )

    total_accounts = 0
    for user_id, _, _, _ in users:
        accounts = get_user_accounts(user_id)
        total_accounts += len(accounts)

    stats_text += f"?? <b>Всего аккаунтов:</b> {total_accounts}\n\n"

    keyboard = InlineKeyboardMarkup(
        inline_keyboard=[
            [
                InlineKeyboardButton(
                    text="?? Скачать список юзеров", callback_data="export_users_excel"
                )
            ]
        ]
    )

    await message.answer(stats_text, parse_mode="HTML", reply_markup=keyboard)


@router.callback_query(F.data == "export_users_excel")
async def export_users_excel(query: CallbackQuery):
    await query.answer("? Создаю список юзеров...", show_alert=False)

    try:
        users = get_all_users()

        wb = Workbook()
        wb.remove(wb.active)
        ws = wb.create_sheet("Юзеры")

        header_fill = PatternFill(
            start_color="4472C4", end_color="4472C4", fill_type="solid"
        )
        header_font = Font(bold=True, color="FFFFFF")

        headers = ["№", "User ID", "Username", "Имя", "VIP", "Статус", "Аккаунтов"]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for row_num, (user_id, username, first_name, is_logged_in) in enumerate(users, 2):
            vip_status = "? VIP" if is_vip_user(user_id) else "?"
            login_status = "? Вход выполнен" if is_logged_in else "? Не вошел"
            account_count = len(get_user_accounts(user_id))

            ws.cell(row=row_num, column=1).value = row_num - 1
            ws.cell(row=row_num, column=2).value = str(user_id)
            ws.cell(row=row_num, column=3).value = username or ""
            ws.cell(row=row_num, column=4).value = first_name or ""
            ws.cell(row=row_num, column=5).value = vip_status
            ws.cell(row=row_num, column=6).value = login_status
            ws.cell(row=row_num, column=7).value = account_count

        ws.column_dimensions["A"].width = 5
        ws.column_dimensions["B"].width = 15
        ws.column_dimensions["C"].width = 20
        ws.column_dimensions["D"].width = 20
        ws.column_dimensions["E"].width = 12
        ws.column_dimensions["F"].width = 20
        ws.column_dimensions["G"].width = 12

        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_bytes = excel_file.getvalue()

        await query.message.answer_document(
            BufferedInputFile(excel_bytes, filename="users_list.xlsx"),
            caption="?? Список всех юзеров бота",
        )
    except Exception as e:
        await query.answer(f"? Ошибка: {str(e)}", show_alert=True)
