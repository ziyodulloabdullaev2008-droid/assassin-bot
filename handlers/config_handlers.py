from aiogram import Router, F

from aiogram.filters.command import Command

from aiogram.fsm.context import FSMContext

from aiogram.fsm.state import State, StatesGroup

from aiogram.types import Message, InlineKeyboardButton, InlineKeyboardMarkup, CallbackQuery, BufferedInputFile

from openpyxl import Workbook

from openpyxl.styles import Font, PatternFill, Alignment

import io

import json
import asyncio



from core.state import app_state

from ui.broadcast_ui import build_broadcast_menu_text

from services.broadcast_profiles_service import (
    list_configs,
    get_active_config_id,
    get_config_detail,
    set_active_config,
    rename_config,
    delete_config,
    export_config_payload,
    import_config_payload,
)


router = Router()





class ConfigRenameState(StatesGroup):

    waiting_name = State()

            



class ConfigUploadState(StatesGroup):

    waiting_file = State()





async def show_config_list(message_or_query, user_id: int, edit: bool = False):

    configs = list_configs(user_id)

    text = "⚙️ <b>КОНФИГИ РАССЫЛКИ</b>\n\nВыберите конфиг:"



    buttons = []

    for cfg_id, name, is_active in configs:

        title = f"✅ {name}" if is_active else name

        buttons.append([InlineKeyboardButton(text=title, callback_data=f"cfg_view_{cfg_id}")])



    buttons.append([

        InlineKeyboardButton(text="⬆️ Загрузить", callback_data="cfg_upload"),

        InlineKeyboardButton(text="⬅️ Назад", callback_data="cfg_close"),

    ])



    kb = InlineKeyboardMarkup(inline_keyboard=buttons)

    if edit:

        await message_or_query.edit_text(text, reply_markup=kb, parse_mode="HTML")

    else:

        await message_or_query.answer(text, reply_markup=kb, parse_mode="HTML")





def _sanitize_filename(name: str) -> str:

    safe = "".join(c for c in name if c.isalnum() or c in "-_ ")

    safe = safe.strip() or "config"

    return safe





async def show_config_detail(query: CallbackQuery, config_id: int):

    user_id = query.from_user.id

    detail = get_config_detail(user_id, config_id)

    if not detail:

        await query.answer("❌ Конфиг не найден", show_alert=True)

        return



    active_id = get_active_config_id(user_id)

    name = detail.get("name", f"Конфиг {config_id}")

    config = detail.get("config", {})

    chats = detail.get("chats", [])



    info = f"⚙️ <b>{name}</b>\n\n"

    info += build_broadcast_menu_text(
        config,
        chats,
        app_state.active_broadcasts,
        user_id,
        show_title=False,
        show_active_count=False,
    )



    buttons = []

    if config_id != active_id:

        buttons.append([InlineKeyboardButton(text="Выбрать конфиг", callback_data=f"cfg_select_{config_id}")])

    else:

        buttons.append([InlineKeyboardButton(text="✅ Текущий", callback_data="cfg_noop")])



    if config_id != 0:

        buttons.append([InlineKeyboardButton(text="Переименовать", callback_data=f"cfg_rename_{config_id}")])



    buttons.append([InlineKeyboardButton(text="Выгрузить чаты", callback_data=f"cfg_export_chats_{config_id}")])

    buttons.append([InlineKeyboardButton(text="Выгрузить конфиг", callback_data=f"cfg_export_{config_id}")])



    if config_id != 0:

        buttons.append([InlineKeyboardButton(text="Удалить конфиг", callback_data=f"cfg_delete_{config_id}")])



    buttons.append([InlineKeyboardButton(text="⬅️ Назад", callback_data="cfg_back")])



    kb = InlineKeyboardMarkup(inline_keyboard=buttons)

    await query.message.edit_text(info, reply_markup=kb, parse_mode="HTML")





@router.message(Command("config"))

async def cmd_config(message: Message):

    await show_config_list(message, message.from_user.id, edit=False)





@router.callback_query(F.data == "cfg_close")

async def cfg_close(query: CallbackQuery):

    await query.answer()

    try:

        await query.message.delete()

    except Exception:

        pass





@router.callback_query(F.data == "cfg_back")

async def cfg_back(query: CallbackQuery):

    await query.answer()

    await show_config_list(query.message, query.from_user.id, edit=True)





@router.callback_query(F.data.startswith("cfg_view_"))

async def cfg_view(query: CallbackQuery):

    await query.answer()

    config_id = int(query.data.split("_")[2])

    if config_id == 0:

        set_active_config(query.from_user.id, 0)

        await show_config_list(query.message, query.from_user.id, edit=True)

        return

    await show_config_detail(query, config_id)





@router.callback_query(F.data.startswith("cfg_select_"))
async def cfg_select(query: CallbackQuery):
    config_id = int(query.data.split("_")[2])
    set_active_config(query.from_user.id, config_id)
    if config_id != 0:
        pass
    await show_config_detail(query, config_id)




@router.callback_query(F.data == "cfg_noop")

async def cfg_noop(query: CallbackQuery):

    await query.answer("Уже выбран", show_alert=False)


@router.callback_query(F.data.startswith("cfg_rename_"))

async def cfg_rename(query: CallbackQuery, state: FSMContext):

    await query.answer()

    config_id = int(query.data.split("_")[2])

    await state.set_state(ConfigRenameState.waiting_name)

    await state.update_data(config_id=config_id)

    await query.message.edit_text("Введите новое название конфига:")





@router.message(ConfigRenameState.waiting_name)

async def cfg_rename_process(message: Message, state: FSMContext):

    data = await state.get_data()

    config_id = data.get("config_id")

    new_name = (message.text or "").strip()

    if not new_name:

        await message.answer("❌ Название не может быть пустым")

        return

    rename_config(message.from_user.id, config_id, new_name)

    await state.clear()

    # Сообщение с вводом названия удаляем, а дальше отправляем новый экран,

    # т.к. редактировать уже нечего.

    try:

        await message.delete()

    except Exception:

        pass

    await show_config_list(message, message.from_user.id, edit=False)





@router.callback_query(F.data.startswith("cfg_delete_"))
async def cfg_delete(query: CallbackQuery):
    await query.answer()
    config_id = int(query.data.split("_")[2])
    delete_config(query.from_user.id, config_id)
    await show_config_list(query.message, query.from_user.id, edit=True)





@router.callback_query(F.data.startswith("cfg_export_chats_"))

async def cfg_export_chats(query: CallbackQuery):

    await query.answer()

    config_id = int(query.data.split("_")[3])

    detail = get_config_detail(query.from_user.id, config_id)

    if not detail:

        await query.answer("❌ Конфиг не найден", show_alert=True)

        return

    chats = detail.get("chats", [])



    wb = Workbook()

    ws = wb.active

    ws.title = "Чаты"



    ws["A1"] = "№"

    ws["B1"] = "Имя"

    ws["C1"] = "ID"



    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    header_font = Font(bold=True, color="FFFFFF")

    for cell in ["A1", "B1", "C1"]:

        ws[cell].fill = header_fill

        ws[cell].font = header_font

        ws[cell].alignment = Alignment(horizontal="center")



    for idx, (chat_id, chat_name) in enumerate(chats, 1):

        ws[f"A{idx+1}"] = idx

        ws[f"B{idx+1}"] = chat_name

        cell = ws[f"C{idx+1}"]
        cell.value = str(chat_id)
        cell.number_format = "@"



    ws.column_dimensions["A"].width = 5

    ws.column_dimensions["B"].width = 40

    ws.column_dimensions["C"].width = 20



    excel_file = io.BytesIO()

    wb.save(excel_file)

    excel_bytes = excel_file.getvalue()



    await query.message.answer_document(

        BufferedInputFile(excel_bytes, filename="broadcast_chats.xlsx"),

        caption="📋 База чатов",

    )





@router.callback_query(F.data.regexp(r"^cfg_export_\d+$"))
async def cfg_export(query: CallbackQuery):
    await query.answer()

    config_id_str = query.data.split("_")[2]

    if not config_id_str.isdigit():

        return

    config_id = int(config_id_str)



    buttons = [

        [InlineKeyboardButton(text="С чатами", callback_data=f"cfg_export_with_{config_id}"),

         InlineKeyboardButton(text="Без чатов", callback_data=f"cfg_export_without_{config_id}")],

        [InlineKeyboardButton(text="⬅️ Назад", callback_data=f"cfg_view_{config_id}")],

    ]

    kb = InlineKeyboardMarkup(inline_keyboard=buttons)

    await query.message.edit_text("Выберите вариант выгрузки:", reply_markup=kb)





async def _send_config_file(query: CallbackQuery, config_id: int, include_chats: bool):

    payload = export_config_payload(query.from_user.id, config_id, include_chats)

    if not payload:

        await query.answer("❌ Конфиг не найден", show_alert=True)

        return

    name = payload.get("name", "config")

    filename = _sanitize_filename(name) + ".json"

    data = json.dumps(payload, ensure_ascii=False, indent=2).encode("utf-8")

    await query.message.answer_document(

        BufferedInputFile(data, filename=filename),

        caption="📦 Конфиг выгружен",

    )





@router.callback_query(F.data.startswith("cfg_export_with_"))

async def cfg_export_with(query: CallbackQuery):

    await query.answer()

    config_id = int(query.data.split("_")[3])

    await _send_config_file(query, config_id, include_chats=True)





@router.callback_query(F.data.startswith("cfg_export_without_"))

async def cfg_export_without(query: CallbackQuery):

    await query.answer()

    config_id = int(query.data.split("_")[3])

    await _send_config_file(query, config_id, include_chats=False)





@router.callback_query(F.data == "cfg_upload")

async def cfg_upload(query: CallbackQuery, state: FSMContext):

    await query.answer()

    await state.set_state(ConfigUploadState.waiting_file)

    await query.message.edit_text("Отправь файл конфига (.json)")





@router.message(ConfigUploadState.waiting_file)

async def cfg_upload_process(message: Message, state: FSMContext):

    if not message.document:

        await message.answer("❌ Отправь .json файл")

        return



    try:

        file = await message.bot.get_file(message.document.file_id)

        file_bytes = await message.bot.download_file(file.file_path)

        payload = json.loads(file_bytes.read().decode("utf-8"))

    except Exception:

        await message.answer("❌ Не удалось прочитать файл")

        return



    new_id = import_config_payload(message.from_user.id, payload)
    await state.clear()

    if new_id is None:
        await message.answer("❌ Неверный формат конфига")
        return

    set_active_config(message.from_user.id, new_id)
    await message.answer("✅ Конфиг загружен и применен")
    await show_config_list(message, message.from_user.id, edit=False)


