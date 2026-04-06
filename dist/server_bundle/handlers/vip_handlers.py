from aiogram import Router, F
from aiogram.filters.command import Command
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.types import (
    Message,
    InlineKeyboardButton,
    InlineKeyboardMarkup,
    CallbackQuery,
    BufferedInputFile,
)
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import io

from core.config import ADMIN_ID
from database import (
    add_vip_user,
    remove_vip_user,
    get_all_vip_users,
    get_all_users,
    get_user_accounts,
    is_vip_user,
)
from services.vip_service import vip_users_cache


def get_vip_cache_size() -> int:
    return len(vip_users_cache)


router = Router()


class VIPAddState(StatesGroup):
    waiting_for_user_id = State()


@router.message(Command("vip"))
async def cmd_add_vip(message: Message):
    """Добавить юзера в VIP по ID: /vip <id>"""
    if message.from_user.id != ADMIN_ID:
        await message.answer("❌ Только владелец может управлять VIP")
        return

    args = message.text.split()
    if len(args) < 2:
        await message.answer("❌ Использование: /vip <user_id>")
        return

    try:
        user_id = int(args[1])
        if add_vip_user(user_id):
            vip_users_cache.add(user_id)
            await message.answer(f"✅ Юзер {user_id} добавлен в VIP")
        else:
            await message.answer(f"⚠️ Юзер {user_id} уже в VIP")
    except ValueError:
        await message.answer("❌ Неверный ID")


@router.message(Command("dlvip"))
async def cmd_remove_vip(message: Message):
    """Удалить юзера из VIP по ID: /dlvip <id>"""
    if message.from_user.id != ADMIN_ID:
        await message.answer("❌ Только владелец может управлять VIP")
        return

    args = message.text.split()
    if len(args) < 2:
        await message.answer("❌ Использование: /dlvip <user_id>")
        return

    try:
        user_id = int(args[1])
        if remove_vip_user(user_id):
            vip_users_cache.discard(user_id)
            await message.answer(f"✅ Юзер {user_id} удален из VIP")
        else:
            await message.answer(f"❌ Юзер {user_id} не найден в VIP")
    except ValueError:
        await message.answer("❌ Неверный ID")


@router.message(Command("vips"))
async def cmd_show_vips(message: Message):
    """Показать список всех VIP юзеров"""
    if message.from_user.id != ADMIN_ID:
        await message.answer("❌ Только владелец может управлять VIP")
        return

    vip_list = get_all_vip_users()
    if not vip_list:
        await message.answer("📭 VIP список пуст")
        return

    text = "👑 <b>VIP Юзеры:</b> ({})\n\n".format(len(vip_list))
    for idx, user_id in enumerate(vip_list, 1):
        text += f"{idx}. <code>{user_id}</code>\n"

    keyboard = InlineKeyboardMarkup(
        inline_keyboard=[
            [InlineKeyboardButton(text="➕ Добавить", callback_data="vip_add_menu")],
            [InlineKeyboardButton(text="❌ Удалить", callback_data="vip_delete_menu")],
        ]
    )

    await message.answer(text, parse_mode="HTML", reply_markup=keyboard)


@router.callback_query(F.data == "vip_add_menu")
async def vip_add_menu(query: CallbackQuery, state: FSMContext):
    """Меню для добавления VIP"""
    if query.from_user.id != ADMIN_ID:
        await query.answer("Только владелец может управлять VIP", show_alert=True)
        return

    await query.answer()
    await state.set_state(VIPAddState.waiting_for_user_id)
    await query.message.edit_text(
        "📝 <b>Введите ID пользователя для добавления:</b>\n(отправьте число)",
        parse_mode="HTML",
    )


@router.callback_query(F.data == "vip_delete_menu")
async def vip_delete_menu(query: CallbackQuery):
    """Меню для удаления VIP"""
    if query.from_user.id != ADMIN_ID:
        await query.answer("Только владелец может управлять VIP", show_alert=True)
        return

    vip_list = get_all_vip_users()
    if not vip_list:
        await query.answer("VIP список пуст", show_alert=True)
        return

    keyboard = InlineKeyboardMarkup(
        inline_keyboard=[
            [InlineKeyboardButton(text=f"❌ {uid}", callback_data=f"vip_remove_{uid}")]
            for uid in vip_list
        ]
        + [[InlineKeyboardButton(text="⬅️ Назад", callback_data="vip_back")]]
    )

    await query.answer()
    await query.message.edit_text(
        "👑 <b>Выберите VIP для удаления:</b>",
        parse_mode="HTML",
        reply_markup=keyboard,
    )


@router.callback_query(F.data.startswith("vip_remove_"))
async def vip_remove_callback(query: CallbackQuery):
    """Удалить VIP из меню"""
    if query.from_user.id != ADMIN_ID:
        await query.answer("Только владелец может управлять VIP", show_alert=True)
        return

    user_id = int(query.data.split("_")[-1])
    if remove_vip_user(user_id):
        vip_users_cache.discard(user_id)
        await query.answer(f"✅ Юзер {user_id} удален из VIP")
        await vip_delete_menu(query)
    else:
        await query.answer("❌ Ошибка удаления", show_alert=True)


@router.callback_query(F.data == "vip_back")
async def vip_back(query: CallbackQuery):
    """Вернуться к списку VIP"""
    if query.from_user.id != ADMIN_ID:
        await query.answer("Только владелец может управлять VIP", show_alert=True)
        return

    vip_list = get_all_vip_users()
    if not vip_list:
        await query.message.edit_text("📭 VIP список пуст")
        return

    text = "👑 <b>VIP Юзеры:</b> ({})\n\n".format(len(vip_list))
    for idx, user_id in enumerate(vip_list, 1):
        text += f"{idx}. <code>{user_id}</code>\n"

    keyboard = InlineKeyboardMarkup(
        inline_keyboard=[
            [InlineKeyboardButton(text="➕ Добавить", callback_data="vip_add_menu")],
            [InlineKeyboardButton(text="❌ Удалить", callback_data="vip_delete_menu")],
        ]
    )

    await query.answer()
    await query.message.edit_text(text, parse_mode="HTML", reply_markup=keyboard)


@router.message(Command("users"))
async def cmd_show_users_stats(message: Message):
    """Показать статистику бота и экспортировать список юзеров"""
    if message.from_user.id != ADMIN_ID:
        await message.answer("❌ Только владелец может использовать эту команду")
        return

    users = get_all_users()
    vip_list = get_all_vip_users()
    vip_count = len(vip_list)

    stats_text = (
        "📊 <b>Статистика бота:</b>\n\n"
        f"👥 <b>Всего юзеров:</b> {len(users)}\n"
        f"👑 <b>VIP юзеров:</b> {vip_count}\n"
        f"📱 <b>Обычных юзеров:</b> {len(users) - vip_count}\n"
    )

    total_accounts = 0
    for user_id, _, _, _ in users:
        accounts = get_user_accounts(user_id)
        total_accounts += len(accounts)

    stats_text += f"🔐 <b>Всего аккаунтов:</b> {total_accounts}\n\n"

    keyboard = InlineKeyboardMarkup(
        inline_keyboard=[
            [
                InlineKeyboardButton(
                    text="📥 Скачать список юзеров", callback_data="export_users_excel"
                )
            ]
        ]
    )

    await message.answer(stats_text, parse_mode="HTML", reply_markup=keyboard)


@router.callback_query(F.data == "export_users_excel")
async def export_users_excel(query: CallbackQuery):
    """Экспортировать список юзеров в Excel"""
    await query.answer("⏳ Создаю список юзеров...", show_alert=False)

    try:
        users = get_all_users()

        wb = Workbook()
        wb.remove(wb.active)
        ws = wb.create_sheet("Юзеры")

        header_fill = PatternFill(
            start_color="4472C4", end_color="4472C4", fill_type="solid"
        )
        header_font = Font(bold=True, color="FFFFFF")

        headers = ["№", "User ID", "Username", "Имя", "VIP", "Статус", "Аккаунтов"]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for row_num, (user_id, username, first_name, is_logged_in) in enumerate(
            users, 2
        ):
            vip_status = "✅ VIP" if is_vip_user(user_id) else "❌"
            login_status = "✅ Вход выполнен" if is_logged_in else "❌ Не вошел"
            accounts = get_user_accounts(user_id)
            account_count = len(accounts)

            ws.cell(row=row_num, column=1).value = row_num - 1
            ws.cell(row=row_num, column=2).value = str(user_id)
            ws.cell(row=row_num, column=3).value = username or ""
            ws.cell(row=row_num, column=4).value = first_name or ""
            ws.cell(row=row_num, column=5).value = vip_status
            ws.cell(row=row_num, column=6).value = login_status
            ws.cell(row=row_num, column=7).value = account_count

        ws.column_dimensions["A"].width = 5
        ws.column_dimensions["B"].width = 15
        ws.column_dimensions["C"].width = 20
        ws.column_dimensions["D"].width = 20
        ws.column_dimensions["E"].width = 12
        ws.column_dimensions["F"].width = 20
        ws.column_dimensions["G"].width = 12

        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_bytes = excel_file.getvalue()

        await query.message.answer_document(
            BufferedInputFile(excel_bytes, filename="users_list.xlsx"),
            caption="📊 Список всех юзеров бота",
        )
    except Exception as e:
        await query.answer(f"❌ Ошибка: {str(e)}", show_alert=True)


@router.message(VIPAddState.waiting_for_user_id)
async def process_vip_user_id(message: Message, state: FSMContext):
    """Обработка ввода ID для добавления в VIP"""
    if message.from_user.id != ADMIN_ID:
        await message.answer("❌ Только владелец может это делать")
        await state.clear()
        return

    try:
        user_id = int(message.text.strip())
        if add_vip_user(user_id):
            vip_users_cache.add(user_id)
            await message.answer(f"✅ Юзер {user_id} добавлен в VIP")
        else:
            await message.answer(f"⚠️ Юзер {user_id} уже в VIP")
    except ValueError:
        await message.answer("❌ Неверный ID. Пожалуйста введите число")
        return
    finally:
        await state.clear()
